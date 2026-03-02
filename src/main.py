"""主入口模块 - CLI 和 Web 启动"""
import argparse
import json
import sys
from pathlib import Path
from typing import Optional

from rich.console import Console
from rich.table import Table
from rich.panel import Panel
from rich import print as rprint

from .config import get_api_key
from .chains import OCRChain, ExtractionChain, ClassificationChain
from .utils import validate_receipt
from .web import launch_app


console = Console()


def recognize_single(
    file_path: str,
    output: Optional[str] = None,
    enable_seal: bool = True,
    verbose: bool = False
) -> dict:
    """识别单张票据

    Args:
        file_path: 图片路径
        output: 输出文件路径
        enable_seal: 是否启用印章识别
        verbose: 是否显示详细信息

    Returns:
        识别结果
    """
    with console.status("[bold green]正在识别票据...") as status:
        # OCR 识别
        status.update("[bold blue]正在进行 OCR 识别...")
        ocr_chain = OCRChain(get_api_key())
        ocr_result = ocr_chain.process(file_path, enable_seal)

        # 信息提取
        status.update("[bold blue]正在提取信息...")
        extraction_chain = ExtractionChain(get_api_key())
        extracted_info = extraction_chain.extract_with_seals(
            ocr_result["text"],
            ocr_result["seals"]
        )

        # 分类
        status.update("[bold blue]正在分类...")
        classification_chain = ClassificationChain(get_api_key())
        classification = classification_chain.classify(ocr_result["text"])
        extracted_info["reimbursement_category"] = classification["reimbursement_category"]

        # 验证
        validation = validate_receipt(extracted_info)
        extracted_info["validation_result"] = validation

    # 显示结果
    display_result(extracted_info, ocr_result["seals"], verbose)

    # 保存结果
    if output:
        save_result(extracted_info, output)

    return extracted_info


def recognize_batch(
    directory: str,
    output: Optional[str] = None,
    enable_seal: bool = True
) -> list:
    """批量识别票据

    Args:
        directory: 图片目录
        output: 输出目录
        enable_seal: 是否启用印章识别

    Returns:
        识别结果列表
    """
    dir_path = Path(directory)
    if not dir_path.is_dir():
        console.print(f"[red]错误：{directory} 不是有效目录[/red]")
        return []

    # 使用 OCR 模块定义的支持格式
    from .ocr.paddle_ocr import PaddleOCRVL
    supported_extensions = PaddleOCRVL.SUPPORTED_EXTENSIONS

    files = [
        f for f in dir_path.iterdir()
        if f.suffix.lower() in supported_extensions and not f.name.startswith(".")
    ]

    if not files:
        console.print(f"[yellow]警告：目录 {directory} 中没有找到支持的文件[/yellow]")
        console.print(f"[yellow]支持的格式：{', '.join(sorted(supported_extensions))}[/yellow]")
        return []

    console.print(f"[bold]找到 {len(files)} 个文件[/bold]")

    results = []
    api_key = get_api_key()
    ocr_chain = OCRChain(api_key)
    extraction_chain = ExtractionChain(api_key)
    classification_chain = ClassificationChain(api_key)

    for i, file_path in enumerate(files, 1):
        console.print(f"\n[bold blue][{i}/{len(files)}][/bold blue] 处理: {file_path.name}")

        try:
            # OCR
            ocr_result = ocr_chain.process(str(file_path), enable_seal)

            # 提取
            info = extraction_chain.extract_with_seals(
                ocr_result["text"],
                ocr_result["seals"]
            )

            # 分类
            classification = classification_chain.classify(ocr_result["text"])
            info["reimbursement_category"] = classification["reimbursement_category"]

            # 验证
            validation = validate_receipt(info)
            info["validation_result"] = validation
            info["file_path"] = str(file_path)

            results.append(info)

            # 显示简要结果
            status = "[green]通过[/green]" if validation.get("is_valid") else "[yellow]需核实[/yellow]"
            console.print(f"  类型: {info.get('receipt_type', '未知')}")
            console.print(f"  金额: {info.get('total', '未知')}")
            console.print(f"  验证: {status}")

        except Exception as e:
            console.print(f"  [red]错误: {str(e)}[/red]")
            results.append({
                "file_path": str(file_path),
                "error": str(e)
            })

    # 保存结果
    if output:
        output_path = Path(output)
        output_path.mkdir(parents=True, exist_ok=True)

        # 保存 JSON
        json_path = output_path / "results.json"
        with open(json_path, "w", encoding="utf-8") as f:
            json.dump(results, f, ensure_ascii=False, indent=2)
        console.print(f"\n[bold green]结果已保存到: {json_path}[/bold green]")

        # 保存 Excel
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "识别结果"

            headers = ["文件名", "类型", "日期", "金额", "销售方", "印章数", "状态"]
            ws.append(headers)

            for cell in ws[1]:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")

            for r in results:
                validation = r.get("validation_result", {})
                ws.append([
                    Path(r.get("file_path", "")).name,
                    r.get("receipt_type", ""),
                    r.get("date", ""),
                    r.get("total", ""),
                    r.get("seller_name", ""),
                    len(r.get("seals", [])) if r.get("seals") else 0,
                    "通过" if validation.get("is_valid") else "需核实"
                ])

            excel_path = output_path / "results.xlsx"
            wb.save(excel_path)
            console.print(f"[bold green]Excel 已保存到: {excel_path}[/bold green]")

        except ImportError:
            console.print("[yellow]提示: 安装 openpyxl 可导出 Excel 格式[/yellow]")

    # 显示统计
    display_batch_summary(results)

    return results


def display_result(info: dict, seals: list, verbose: bool = False):
    """显示识别结果

    Args:
        info: 提取的信息
        seals: 印章列表
        verbose: 是否显示详细信息
    """
    # 基本信息
    table = Table(title="票据信息")
    table.add_column("字段", style="cyan")
    table.add_column("值", style="green")

    fields = [
        ("票据类型", info.get("receipt_type")),
        ("发票代码", info.get("invoice_code")),
        ("发票号码", info.get("invoice_number")),
        ("开票日期", info.get("date")),
        ("购买方", info.get("buyer_name")),
        ("销售方", info.get("seller_name")),
        ("金额", info.get("amount")),
        ("税额", info.get("tax")),
        ("价税合计", info.get("total")),
        ("报销类别", info.get("reimbursement_category")),
    ]

    for name, value in fields:
        if value:
            table.add_row(name, str(value))

    console.print(table)

    # 印章信息
    if seals:
        seal_table = Table(title="印章信息")
        seal_table.add_column("类型", style="cyan")
        seal_table.add_column("状态", style="green")

        for seal in seals:
            seal_table.add_row(seal.get("type", "未知"), "已识别")

        console.print(seal_table)

        if info.get("seal_analysis"):
            hint = info["seal_analysis"].get("authenticity_hint", "")
            console.print(f"\n[bold]真伪提示:[/bold] {hint}")

    # 验证结果
    validation = info.get("validation_result", {})
    if validation:
        if validation.get("is_valid"):
            console.print("\n[bold green]验证通过[/bold green]")
        else:
            console.print("\n[bold red]验证失败[/bold red]")
            for issue in validation.get("issues", []):
                console.print(f"  - {issue}")

        for warning in validation.get("warnings", []):
            console.print(f"  [yellow]警告: {warning}[/yellow]")


def display_batch_summary(results: list):
    """显示批量处理统计

    Args:
        results: 结果列表
    """
    console.print("\n" + "=" * 50)

    # 统计
    total = len(results)
    success = sum(1 for r in results if not r.get("error"))
    failed = total - success

    # 类型分布
    type_counts = {}
    category_counts = {}
    total_amount = 0

    for r in results:
        if not r.get("error"):
            t = r.get("receipt_type", "未知")
            type_counts[t] = type_counts.get(t, 0) + 1

            c = r.get("reimbursement_category", "其他")
            category_counts[c] = category_counts.get(c, 0) + 1

            try:
                amount = float(r.get("total", "0").replace(",", ""))
                total_amount += amount
            except:
                pass

    # 显示统计
    summary_table = Table(title="处理统计")
    summary_table.add_column("项目", style="cyan")
    summary_table.add_column("数量", style="green")

    summary_table.add_row("总数", str(total))
    summary_table.add_row("成功", str(success))
    summary_table.add_row("失败", str(failed))
    summary_table.add_row("总金额", f"¥{total_amount:,.2f}")

    console.print(summary_table)

    # 类型分布
    if type_counts:
        type_table = Table(title="票据类型分布")
        type_table.add_column("类型", style="cyan")
        type_table.add_column("数量", style="green")

        for t, count in type_counts.items():
            type_table.add_row(t, str(count))

        console.print(type_table)

    # 报销类别分布
    if category_counts:
        cat_table = Table(title="报销类别分布")
        cat_table.add_column("类别", style="cyan")
        cat_table.add_column("数量", style="green")

        for c, count in category_counts.items():
            cat_table.add_row(c, str(count))

        console.print(cat_table)


def save_result(result: dict, output: str):
    """保存结果到文件

    Args:
        result: 结果字典
        output: 输出路径
    """
    output_path = Path(output)

    # 根据扩展名决定格式
    if output_path.suffix.lower() == ".json":
        with open(output_path, "w", encoding="utf-8") as f:
            json.dump(result, f, ensure_ascii=False, indent=2)
    else:
        # 默认 JSON
        with open(str(output_path) + ".json", "w", encoding="utf-8") as f:
            json.dump(result, f, ensure_ascii=False, indent=2)

    console.print(f"[bold green]结果已保存到: {output_path}[/bold green]")


def cli():
    """CLI 入口"""
    parser = argparse.ArgumentParser(
        description="智能票据报销助手 - PaddleOCR + LangChain + ERINE",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
示例:
  # 识别单张票据
  uv run python -m src.main recognize invoice.jpg

  # 识别并保存结果
  uv run python -m src.main recognize invoice.jpg --output result.json

  # 批量识别
  uv run python -m src.main batch ./invoices/ --output ./results/

  # 启动 Web 界面
  uv run python -m src.main web
        """
    )

    subparsers = parser.add_subparsers(dest="command", help="命令")

    # recognize 命令
    recognize_parser = subparsers.add_parser("recognize", help="识别单张票据")
    recognize_parser.add_argument("file", help="票据图片路径")
    recognize_parser.add_argument("--output", "-o", help="输出文件路径")
    recognize_parser.add_argument("--no-seal", action="store_true", help="禁用印章识别")
    recognize_parser.add_argument("--verbose", "-v", action="store_true", help="显示详细信息")

    # batch 命令
    batch_parser = subparsers.add_parser("batch", help="批量识别票据")
    batch_parser.add_argument("directory", help="图片目录")
    batch_parser.add_argument("--output", "-o", help="输出目录")
    batch_parser.add_argument("--no-seal", action="store_true", help="禁用印章识别")

    # web 命令
    web_parser = subparsers.add_parser("web", help="启动 Web 界面")
    web_parser.add_argument("--share", action="store_true", help="创建公共链接")
    web_parser.add_argument("--port", type=int, default=7860, help="端口号")
    web_parser.add_argument("--host", default="0.0.0.0", help="主机地址")

    args = parser.parse_args()

    if args.command == "recognize":
        recognize_single(
            args.file,
            args.output,
            enable_seal=not args.no_seal,
            verbose=args.verbose
        )

    elif args.command == "batch":
        recognize_batch(
            args.directory,
            args.output,
            enable_seal=not args.no_seal
        )

    elif args.command == "web":
        console.print(Panel.fit(
            "[bold green]智能票据报销助手[/bold green]\n"
            "基于 PaddleOCR-VL-1.5 + LangChain + ERINE",
            border_style="green"
        ))
        launch_app(share=args.share, server_name=args.host, server_port=args.port)

    else:
        parser.print_help()


if __name__ == "__main__":
    cli()
