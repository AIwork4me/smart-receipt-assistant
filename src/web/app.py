"""Gradio Web 应用"""
import gradio as gr
import json
import os
from typing import Optional, List
from datetime import datetime
from pathlib import Path

from ..config import get_api_key
from ..chains import OCRChain, ExtractionChain, ClassificationChain
from ..utils import validate_receipt
from .components import (
    create_upload_component,
    create_result_component,
    create_batch_component,
    create_sample_component,
    format_seal_info,
    format_validation_result,
    get_sample_path,
    SAMPLE_INVOICES
)


def process_single_receipt(
    file_path: str,
    enable_seal: bool = True
) -> tuple:
    """处理单张票据

    Args:
        file_path: 文件路径（支持 PDF 和图片）
        enable_seal: 是否启用印章识别

    Returns:
        (json_result, text_result, seal_images, seal_info, validation_status)
    """
    if not file_path:
        return None, "", [], "请上传文件", ""

    try:
        api_key = get_api_key()
        # OCR 识别
        ocr_chain = OCRChain(api_key)
        ocr_result = ocr_chain.process(file_path, enable_seal)

        # 信息提取
        extraction_chain = ExtractionChain(api_key)
        extracted_info = extraction_chain.extract_with_seals(
            ocr_result["text"],
            ocr_result["seals"]
        )

        # 分类
        classification_chain = ClassificationChain(api_key)
        classification = classification_chain.classify(ocr_result["text"])
        extracted_info["reimbursement_category"] = classification["reimbursement_category"]

        # 验证
        validation = validate_receipt(extracted_info)
        extracted_info["validation_result"] = validation

        # 格式化印章信息
        seal_images, seal_info_text = format_seal_info(
            ocr_result["seals"],
            extracted_info.get("seal_analysis")
        )

        # 格式化验证结果
        validation_text = format_validation_result(validation)

        return (
            extracted_info,
            ocr_result["text"],
            seal_images,
            seal_info_text,
            validation_text
        )

    except Exception as e:
        return None, "", [], f"处理出错：{str(e)}", ""


def process_sample(sample_file: str, enable_seal: bool = True) -> tuple:
    """处理样本发票

    Args:
        sample_file: 样本文件名
        enable_seal: 是否启用印章识别

    Returns:
        (file_path, json_result, text_result, seal_images, seal_info, validation_status, sample_info)
    """
    try:
        file_path = get_sample_path(sample_file)

        # 获取样本信息
        sample_info = next(
            (s for s in SAMPLE_INVOICES if s["file"] == sample_file),
            {"name": sample_file, "description": ""}
        )
        info_text = f"类型: {sample_info.get('type', '未知')} | {sample_info.get('description', '')}"

        # 处理发票
        result = process_single_receipt(file_path, enable_seal)

        return (file_path,) + result + (info_text,)

    except Exception as e:
        return None, None, "", [], f"处理出错：{str(e)}", "", ""


def process_batch_receipts(
    files: List[str],
    enable_seal: bool = True
) -> tuple:
    """批量处理票据

    Args:
        files: 文件路径列表
        enable_seal: 是否启用印章识别

    Returns:
        (results_dataframe, export_file)
    """
    if not files:
        return [], None

    results = []
    api_key = get_api_key()
    ocr_chain = OCRChain(api_key)
    extraction_chain = ExtractionChain(api_key)
    classification_chain = ClassificationChain(api_key)

    for file_path in files:
        try:
            # OCR
            ocr_result = ocr_chain.process(file_path, enable_seal)

            # 提取
            info = extraction_chain.extract_with_seals(
                ocr_result["text"],
                ocr_result["seals"]
            )

            # 分类
            classification = classification_chain.classify(ocr_result["text"])
            info["reimbursement_category"] = classification["reimbursement_category"]

            # 验证
            validation = validate_receipt(info)

            results.append([
                Path(file_path).name,
                info.get("receipt_type", ""),
                info.get("date", ""),
                info.get("total", ""),
                info.get("seller_name", ""),
                len(ocr_result["seals"]) if ocr_result["seals"] else 0,
                "通过" if validation.get("is_valid") else "需核实"
            ])

        except Exception as e:
            results.append([
                Path(file_path).name,
                "",
                "",
                "",
                "",
                0,
                f"错误: {str(e)[:20]}"
            ])

    return results, None


def export_to_json(result: dict) -> str:
    """导出为 JSON 文件

    Args:
        result: 识别结果

    Returns:
        文件路径
    """
    if not result:
        return None

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"receipt_{timestamp}.json"

    with open(filename, "w", encoding="utf-8") as f:
        json.dump(result, f, ensure_ascii=False, indent=2)

    return filename


def export_to_excel(results: list) -> str:
    """导出为 Excel 文件

    Args:
        results: 结果列表

    Returns:
        文件路径
    """
    if not results:
        return None

    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "票据识别结果"

        # 表头
        headers = ["文件名", "票据类型", "日期", "金额", "销售方", "印章数量", "状态"]
        ws.append(headers)

        # 设置表头样式
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")

        # 数据
        for row in results:
            ws.append(row)

        # 调整列宽
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            ws.column_dimensions[column].width = min(max_length + 2, 50)

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"receipts_{timestamp}.xlsx"
        wb.save(filename)

        return filename

    except ImportError:
        # 如果没有 openpyxl，导出为 CSV
        import csv
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"receipts_{timestamp}.csv"

        with open(filename, "w", newline="", encoding="utf-8-sig") as f:
            writer = csv.writer(f)
            writer.writerow(["文件名", "票据类型", "日期", "金额", "销售方", "印章数量", "状态"])
            writer.writerows(results)

        return filename


def create_app() -> gr.Blocks:
    """创建 Gradio 应用

    Returns:
        Gradio Blocks 应用
    """
    with gr.Blocks(title="智能票据报销助手") as app:
        gr.Markdown(
            """
            # 智能票据报销助手
            基于 PaddleOCR-VL-1.5 + LangChain + ERINE 的票据识别与信息提取系统

            **特色功能**：支持印章识别，辅助验证发票真伪
            """
        )

        with gr.Tabs():
            # 单张识别
            with gr.TabItem("单张识别"):
                with gr.Row():
                    # 左侧：上传区域 + 样本发票
                    with gr.Column(scale=1):
                        upload_col, file_input, enable_seal, submit_btn = create_upload_component()

                        gr.Markdown("---")

                        sample_col, sample_buttons, sample_info = create_sample_component()

                    # 右侧：结果展示
                    with gr.Column(scale=2):
                        right_col, result_json, result_text, seal_gallery, seal_info, validation_status = create_result_component()

                # 绑定上传识别按钮
                submit_btn.click(
                    fn=process_single_receipt,
                    inputs=[file_input, enable_seal],
                    outputs=[result_json, result_text, seal_gallery, seal_info, validation_status]
                )

                # 绑定样本按钮
                for btn, sample in sample_buttons:
                    btn.click(
                        fn=lambda s=sample["file"], e=enable_seal: process_sample(s, e),
                        inputs=[],
                        outputs=[file_input, result_json, result_text, seal_gallery, seal_info, validation_status, sample_info]
                    )

            # 批量处理
            with gr.TabItem("批量处理"):
                batch_col, batch_file_input, batch_btn, batch_results, batch_export_btn, batch_file = create_batch_component()

                batch_btn.click(
                    fn=process_batch_receipts,
                    inputs=[batch_file_input],
                    outputs=[batch_results, batch_file]
                )

                batch_export_btn.click(
                    fn=export_to_excel,
                    inputs=[batch_results],
                    outputs=[batch_file]
                )

        gr.Markdown(
            """
            ---
            **使用说明**：
            1. 上传票据文件（支持 PDF 和图片格式：jpg, png, bmp, tiff, webp）
            2. 或点击左侧样本发票按钮，快速体验识别效果
            3. 查看识别结果和印章信息
            4. 可导出 JSON 或 Excel 格式

            **支持的票据类型**：增值税发票、火车票、出租车票等
            """
        )

    return app


def launch_app(share: bool = False, server_name: str = "0.0.0.0", server_port: int = 7860):
    """启动 Gradio 应用

    Args:
        share: 是否创建公共链接
        server_name: 服务器名称
        server_port: 端口号
    """
    app = create_app()
    app.launch(
        share=share,
        server_name=server_name,
        server_port=server_port
    )


if __name__ == "__main__":
    launch_app()
